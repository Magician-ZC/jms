"""
虚假签收报表模块
获取虚假签收数据并导出Excel
"""
import json
import requests
from datetime import datetime, timedelta
from typing import Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

from config import FALSE_SIGN_REPORT_API, FALSE_SIGN_EXCEL_COLUMNS


class FalseSignReport:
    """虚假签收报表爬取器"""

    def __init__(self, authtoken: Optional[str] = None):
        self.authtoken = authtoken or self._load_authtoken()
        self.headers = self._build_headers()

    def _load_authtoken(self) -> str:
        """从文件加载authtoken"""
        try:
            with open("authtoken.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("authtoken", "")
        except:
            return ""

    def _build_headers(self) -> dict:
        """构建请求头"""
        return {
            "accept": "application/json, text/plain, */*",
            "authtoken": self.authtoken,
            "content-type": "application/json;charset=UTF-8",
            "lang": "zh_CN",
            "origin": "https://idata.jtexpress.com.cn",
            "referer": "https://idata.jtexpress.com.cn/",
            "routename": "FalseSignReportPC|serviceQualityIndex",
        }

    def _build_request_body(self, page: int = 1, size: int = 50, date: Optional[str] = None) -> dict:
        """
        构建请求体
        Args:
            page: 页码
            size: 每页数量
            date: 日期，格式 YYYY-MM-DD，默认为昨天
        """
        if date is None:
            # 默认昨天
            target_date = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        else:
            target_date = date

        return {
            "current": page,
            "size": size,
            "workRemarks": ["-"],
            "startTime": f"{target_date} 00:00:00",
            "endTime": f"{target_date} 23:59:59",
            "agentCodes": ["350000"],
            "agentType": 1,
            "countryId": "1",
            "isvirtualAgent": 0,
        }

    def _fetch_page(self, page: int, size: int = 50, date: Optional[str] = None) -> tuple:
        """获取单页数据，返回 (records, total, pages)"""
        body = self._build_request_body(page=page, size=size, date=date)
        try:
            response = requests.post(
                FALSE_SIGN_REPORT_API,
                headers=self.headers,
                data=json.dumps(body, separators=(",", ":")),
                timeout=15,
            )
            data = response.json()
            
            # 调试输出
            if page == 1:
                print(f"[调试] API响应: code={data.get('code')}, succ={data.get('succ')}, msg={data.get('msg')}")
                if not data.get('succ'):
                    print(f"[调试] 完整响应: {json.dumps(data, ensure_ascii=False)[:500]}")
            
            if data.get("code") == 1 and data.get("succ"):
                result = data.get("data", {})
                return (
                    result.get("records", []),
                    result.get("total", 0),
                    result.get("pages", 1),
                )
        except Exception as e:
            print(f"[错误] 获取第{page}页失败: {e}")
        return ([], 0, 1)

    def fetch_all(self, date: Optional[str] = None) -> list:
        """
        获取所有虚假签收数据
        Args:
            date: 日期，格式 YYYY-MM-DD，默认为昨天
        """
        if not self.authtoken:
            print("[错误] 无有效authtoken")
            return []

        target_date = date or (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        print(f"[虚假签收] 开始获取 {target_date} 的数据...")

        # 先获取第一页，确定总页数
        records, total, total_pages = self._fetch_page(1, size=50, date=target_date)
        if not records:
            print("[虚假签收] 无数据")
            return []

        all_records = list(records)
        print(f"[虚假签收] 总记录数: {total}, 总页数: {total_pages}")

        # 并行获取剩余页
        if total_pages > 1:
            with ThreadPoolExecutor(max_workers=6) as executor:
                futures = [
                    executor.submit(self._fetch_page, page, 50, target_date)
                    for page in range(2, total_pages + 1)
                ]
                for future in as_completed(futures):
                    page_records, _, _ = future.result()
                    all_records.extend(page_records)

        print(f"[虚假签收] 共获取 {len(all_records)} 条记录")
        return all_records

    def export_excel(self, date: Optional[str] = None, output_path: Optional[str] = None) -> str:
        """
        导出虚假签收报表到Excel
        Args:
            date: 日期，格式 YYYY-MM-DD，默认为昨天
            output_path: 输出文件路径，默认为 虚假签收报表_日期.xlsx
        Returns:
            导出的文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            print("[错误] 请先安装 openpyxl: pip install openpyxl")
            return ""

        # 获取数据
        records = self.fetch_all(date=date)
        if not records:
            print("[导出] 无数据可导出")
            return ""

        target_date = date or (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        if output_path is None:
            output_path = f"虚假签收报表_{target_date}.xlsx"

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "虚假签收报表"

        # 样式定义
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 写入表头
        columns = list(FALSE_SIGN_EXCEL_COLUMNS.keys())
        headers = list(FALSE_SIGN_EXCEL_COLUMNS.values())

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # 写入数据
        for row_idx, record in enumerate(records, 2):
            for col_idx, field in enumerate(columns, 1):
                value = record.get(field, "")
                # 清理字符串中的空格
                if isinstance(value, str):
                    value = value.strip()
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 调整列宽
        column_widths = {
            "工单号": 20,
            "运单号": 20,
            "工单创建时间": 20,
            "工单类型": 12,
            "投诉对象": 10,
            "问题类型": 15,
            "问题描述": 40,
            "虚假类型": 25,
            "签收类型": 12,
            "代理区": 12,
            "虚拟网点": 12,
            "片区": 10,
            "网点名称": 18,
            "网点编码": 12,
            "派件员": 15,
            "派件员编码": 12,
            "订单来源": 12,
        }
        for col_idx, header in enumerate(headers, 1):
            width = column_widths.get(header, 10)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        # 保存
        wb.save(output_path)
        print(f"[导出成功] {output_path} (共 {len(records)} 条记录)")
        return output_path


def download_false_sign_report(date: Optional[str] = None, output_path: Optional[str] = None) -> str:
    """
    下载虚假签收报表（对外接口）
    Args:
        date: 日期，格式 YYYY-MM-DD，默认为昨天
        output_path: 输出文件路径
    Returns:
        导出的文件路径
    """
    report = FalseSignReport()
    return report.export_excel(date=date, output_path=output_path)


if __name__ == "__main__":
    # 测试：下载昨天的虚假签收报表
    download_false_sign_report()
