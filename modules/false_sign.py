"""
虚假签收报表模块
支持代理区(agent)和网点(network)两种账号类型
"""
import json
import requests
from datetime import datetime, timedelta
from typing import Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

from config import (
    AccountType,
    AGENT_CONFIG, NETWORK_CONFIG,
    FALSE_SIGN_AGENT_API, FALSE_SIGN_NETWORK_API,
    FALSE_SIGN_AGENT_COLUMNS, FALSE_SIGN_NETWORK_COLUMNS,
)


class FalseSignModule:
    """虚假签收报表工具类，支持代理区和网点两种账号类型"""

    def __init__(self, authtoken: str, account_type: str = "agent"):
        self.authtoken = authtoken
        if isinstance(account_type, str):
            self.account_type = AccountType(account_type.lower())
        else:
            self.account_type = account_type
        self.config = self._get_config()
        self.headers = self._build_headers()

    def _get_config(self) -> dict:
        if self.account_type == AccountType.AGENT:
            return {"api_url": FALSE_SIGN_AGENT_API, "columns": FALSE_SIGN_AGENT_COLUMNS, **AGENT_CONFIG}
        return {"api_url": FALSE_SIGN_NETWORK_API, "columns": FALSE_SIGN_NETWORK_COLUMNS, **NETWORK_CONFIG}

    def _build_headers(self) -> dict:
        headers = {
            "accept": "application/json, text/plain, */*",
            "content-type": "application/json;charset=UTF-8",
            "lang": "zh_CN",
            "origin": self.config["origin"],
            "referer": self.config["referer"],
        }
        headers[self.config["token_header"]] = self.authtoken
        if self.account_type == AccountType.AGENT:
            headers["routename"] = "FalseSignReportPC|serviceQualityIndex"
        else:
            headers["routeName"] = "falseSignReportPC"
        return headers

    def _build_request_body(self, page: int = 1, size: int = 50, date: str = None) -> dict:
        target_date = date or (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        body = {
            "current": page,
            "size": size,
            "workRemarks": ["-"],
            "startTime": f"{target_date} 00:00:00",
            "endTime": f"{target_date} 23:59:59",
            "countryId": "1",
        }
        if self.account_type == AccountType.AGENT:
            body.update({"agentCodes": ["350000"], "agentType": 1, "isvirtualAgent": 0})
        return body

    def _fetch_page(self, page: int, size: int = 50, date: str = None) -> tuple:
        body = self._build_request_body(page=page, size=size, date=date)
        try:
            response = requests.post(
                self.config["api_url"],
                headers=self.headers,
                data=json.dumps(body, separators=(",", ":")),
                timeout=15,
            )
            if page == 1:
                print(f"[调试] HTTP状态码: {response.status_code}")
            if response.status_code == 401:
                print("[错误] 401未授权，token可能已失效")
                return ([], 0, 1)
            data = response.json()
            if page == 1:
                print(f"[调试] API响应: code={data.get('code')}, succ={data.get('succ')}")
            if data.get("code") == 1 and data.get("succ"):
                result = data.get("data", {})
                return (result.get("records", []), result.get("total", 0), result.get("pages", 1))
        except Exception as e:
            print(f"[错误] 获取第{page}页失败: {e}")
        return ([], 0, 1)

    def fetch_all(self, date: str = None) -> list:
        if not self.authtoken:
            print("[错误] 无有效authtoken")
            return []
        target_date = date or (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        type_name = self.config["name"]
        print(f"[虚假签收-{type_name}] 开始获取 {target_date} 的数据...")
        records, total, total_pages = self._fetch_page(1, size=50, date=target_date)
        if not records:
            print(f"[虚假签收-{type_name}] 无数据")
            return []
        all_records = list(records)
        print(f"[虚假签收-{type_name}] 总记录数: {total}, 总页数: {total_pages}")
        if total_pages > 1:
            with ThreadPoolExecutor(max_workers=6) as executor:
                futures = [executor.submit(self._fetch_page, p, 50, target_date) for p in range(2, total_pages + 1)]
                for future in as_completed(futures):
                    page_records, _, _ = future.result()
                    all_records.extend(page_records)
        print(f"[虚假签收-{type_name}] 共获取 {len(all_records)} 条记录")
        return all_records

    def export_excel(self, date: str = None, output_path: str = None) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            print("[错误] 请先安装 openpyxl")
            return ""
        records = self.fetch_all(date=date)
        if not records:
            return ""
        target_date = date or (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        type_suffix = "代理区" if self.account_type == AccountType.AGENT else "网点"
        if output_path is None:
            output_path = f"虚假签收报表_{type_suffix}_{target_date}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "虚假签收报表"
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        excel_columns = self.config["columns"]
        columns = list(excel_columns.keys())
        headers = list(excel_columns.values())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font, cell.fill, cell.border, cell.alignment = header_font, header_fill, thin_border, center_align
        for row_idx, record in enumerate(records, 2):
            for col_idx, field in enumerate(columns, 1):
                value = record.get(field, "")
                if isinstance(value, str):
                    value = value.strip()
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        ws.freeze_panes = "A2"
        wb.save(output_path)
        print(f"[导出成功] {output_path} (共 {len(records)} 条记录)")
        return output_path

    def run(self, date: str = None) -> str:
        type_name = self.config["name"]
        print(f"\n[下载] 开始下载{type_name}虚假签收报表...")
        target_date = date or (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        return self.export_excel(date=target_date)


def create_agent_module(authtoken: str) -> FalseSignModule:
    return FalseSignModule(authtoken=authtoken, account_type="agent")


def create_network_module(authtoken: str) -> FalseSignModule:
    return FalseSignModule(authtoken=authtoken, account_type="network")
